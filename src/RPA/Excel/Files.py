import logging
import pathlib
from contextlib import contextmanager
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

import xlrd
import xlwt
from xlutils.copy import copy as xlutils_copy

from RPA.Tables import Tables


class Files:
    """Robot Framework library for manipulating Excel files.

    Note: To run macros or load password protected worksheets,
    please use the Excel application library.
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.workbook = None

    def _load_workbook(self, path):
        path = pathlib.Path(path).resolve()

        try:
            book = XlsxWorkbook(path)
            book.open()
            return book
        except openpyxl.utils.exceptions.InvalidFileException as err:
            self.logger.debug(err)

        self.logger.info(
            "Failed to open in Office Open XML (.xlsx) format, "
            "trying Excel Binary Format (.xls)"
        )

        try:
            book = XlsWorkbook(path)
            book.open()
            return book
        except xlrd.biffh.XLRDError as err:
            self.logger.debug(err)

        raise ValueError(f"Not a valid Excel file: {path}")

    def create_workbook(self, path=None, fmt="xlsx"):
        """Create and open a new Excel workbook.

        :param path: Default save path for workbook
        :param fmt:  Format of workbook, i.e. xlsx or xls
        """
        if self.workbook:
            self.close_workbook()

        fmt = str(fmt).lower().strip()
        if fmt == "xlsx":
            self.workbook = XlsxWorkbook(path)
        elif fmt == "xls":
            self.workbook = XlsWorkbook(path)
        else:
            raise ValueError(f"Unknown format: {fmt}")

        self.workbook.create()
        return self.workbook

    def open_workbook(self, path):
        """Open an existing Excel workbook.

        :param path: path to Excel file
        """
        if self.workbook:
            self.close_workbook()

        self.workbook = self._load_workbook(path)
        self.logger.info("Opened workbook: %s", self.workbook)
        return self.workbook

    def close_workbook(self):
        """Close the active workbook."""
        if self.workbook:
            self.logger.info("Closing workbook: %s", self.workbook)
            self.workbook.close()

    def save_workbook(self, path=None):
        """Save the active workbook.

        :param path: Path to save to. If not given, uses path given
                     when opened or created.
        """
        assert self.workbook, "No active workbook"
        return self.workbook.save(path)

    def list_worksheets(self):
        """List all names of worksheets in the given workbook."""
        assert self.workbook, "No active workbook"
        return self.workbook.sheetnames

    def worksheet_exists(self, name):
        """Return True if worksheet with given name is in workbook."""
        assert self.workbook, "No active workbook"
        return bool(str(name) in self.list_worksheets())

    def get_active_worksheet(self):
        """Get the name of the worksheet which is currently active."""
        assert self.workbook, "No active workbook"
        return self.workbook.active

    def set_active_worksheet(self, value):
        """Set the active worksheet.

        :param value: Index or name of worksheet
        """
        assert self.workbook, "No active workbook"
        self.workbook.active = value

    def create_worksheet(self, name, content=None, exist_ok=False):
        """Create a new worksheet in the current workbook.

        :param name:     Name of new worksheet
        :param content:  Optional content for worksheet
        :param exist_ok: If `False`, raise an error if name is already in use
        """
        assert self.workbook, "No active workbook"
        if name in self.workbook.sheetnames and not exist_ok:
            raise ValueError(f"Sheet with name {name} already exists")

        self.workbook.create_worksheet(name)
        if content:
            self.workbook.append_worksheet(name, content)

    def read_worksheet(self, name=None, header=False):
        """Read the content of a worksheet into a list of dictionaries.

        Each key in the dictionary will be either values from the header row,
        or Excel-style column letters.

        :param name:   Name of worksheet to read
        :param header: If `True`, use the first row of the worksheet
                       as headers for the rest of the rows.
        """
        assert self.workbook, "No active workbook"
        return self.workbook.read_worksheet(name, header)

    def read_worksheet_as_table(self, name=None, header=False, trim=True):
        """Read the content of a worksheet into a Table container. Allows
        sorting/filtering/manipulating using the `RPA.Tables` library.

        :param name:   Name of worksheet to read
        :param header: If `True`, use the first row of the worksheet
                       as headers for the rest of the rows.
        :param trim:   Remove all empty rows from the end of the worksheet
        """
        library = Tables()
        sheet = self.read_worksheet(name, header)

        table = library.create_table(sheet)
        if trim:
            library.trim_empty_rows(table)

        return table

    def append_rows_to_worksheet(self, content, name=None):
        """Append values to the end of the worksheet.

        :param content: Rows of values to append
        :param name:    Name of worksheet to append to
        """
        assert self.workbook, "No active workbook"
        return self.workbook.append_worksheet(name, content)

    def remove_worksheet(self, name=None):
        """Remove a worksheet from the active workbook.

        :param name: Name of worksheet to remove
        """
        assert self.workbook, "No active workbook"
        self.workbook.remove_worksheet(name)

    def rename_worksheet(self, src_name, dst_name):
        """Rename an existing worksheet in the active workbook.

        :param src_name: Current name of worksheet
        :param dst_name: Future name of worksheet
        """
        assert self.workbook, "No active workbook"
        self.workbook.rename_worksheet(dst_name, src_name)


class XlsxWorkbook:
    """Container for manipulating moden Excel files (.xlsx)"""

    def __init__(self, path=None):
        self.path = path
        self._book = None
        self._active = None

    @property
    def sheetnames(self):
        return list(self._book.sheetnames)

    @property
    def active(self):
        if not self._active:
            self._active = self._book.active.title

        return self._active

    @active.setter
    def active(self, value):
        if isinstance(value, int):
            value = self.sheetnames[value]
        elif value not in self.sheetnames:
            raise ValueError(f"Unknown worksheet: {value}")

        self._book.active = self.sheetnames.index(value)
        self._active = value

    def _get_sheetname(self, name=None):
        if not self.sheetnames:
            raise ValueError("No worksheets in file")

        if name is None:
            name = self.active
        elif isinstance(name, int):
            name = self.sheetnames[name]

        return name

    def _get_columns(self, sheet, header):
        if header:
            columns = [cell.value for cell in sheet[1]]
            start = 2
        else:
            columns = [get_column_letter(i + 1) for i in range(sheet.max_column)]
            start = 1

        return start, columns

    def create(self):
        self._book = openpyxl.Workbook()

    def open(self, path=None):
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")
        self._book = openpyxl.load_workbook(filename=path, keep_vba=True)

    def close(self):
        self._book.close()
        self._book = None
        self._active = None

    def save(self, path=None):
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")
        self._book.save(filename=path)

    def create_worksheet(self, name):
        self._book.create_sheet(title=name)
        self.active = name

    def read_worksheet(self, name=None, header=False):
        name = self._get_sheetname(name)
        sheet = self._book[name]
        start, columns = self._get_columns(sheet, header)

        data = []
        for cells in sheet.iter_rows(min_row=start):
            row = {}
            for c, cell in enumerate(cells):
                column = columns[c]
                if column is not None:
                    row[column] = cell.value
            data.append(row)

        self.active = name
        return data

    def append_worksheet(self, name=None, content=None, header=False):
        content = content or []
        name = self._get_sheetname(name)
        sheet = self._book[name]
        _, columns = self._get_columns(sheet, header)

        for row in content:
            if header and isinstance(row, dict):
                for column, value in row.items():
                    index = columns.index(column)
                    row[index] = value
            sheet.append(row)

        self.active = name

    def remove_worksheet(self, name=None):
        name = self._get_sheetname(name)
        sheet = self._book[name]
        self._book.remove(sheet)

    def rename_worksheet(self, title, name=None):
        title = str(title)
        name = self._get_sheetname(name)
        sheet = self._book[name]

        sheet.title = title
        self.active = title


class XlsWorkbook:
    """Container for manipulating legacy Excel files (.xls)"""

    def __init__(self, path=None):
        self.path = path
        self._book = None
        self._active = None

    @property
    def sheetnames(self):
        return [sheet.name for sheet in self._book.sheets()]

    @property
    def active(self):
        if not self._active:
            for sheet in self._book.sheets():
                if sheet.sheet_visible:
                    self._active = sheet.name
                    break
            else:
                self._active = self.sheetnames[0]

        return self._active

    @active.setter
    def active(self, value):
        if isinstance(value, int):
            value = self.sheetnames[value]
        elif value not in self.sheetnames:
            raise ValueError(f"Unknown worksheet: {value}")

        for sheet in self._book.sheets():
            match = int(sheet.name == value)
            sheet.sheet_selected = match
            sheet.sheet_visible = match

        self._active = value

    def _get_sheetname(self, name):
        if self._book.nsheets == 0:
            raise ValueError("No worksheets in file")

        if name is None:
            name = self.active
        elif isinstance(name, int):
            name = self.sheetnames[name]
        elif name not in self.sheetnames:
            raise ValueError(f"Unknown worksheet: {name}")

        return name

    def _get_columns(self, sheet, header):
        if header:
            columns = [cell.value for cell in sheet.row(0)]
            start = 1
        else:
            columns = [get_column_letter(i + 1) for i in range(sheet.ncols)]
            start = 0

        return start, columns

    def create(self, sheet="Sheet"):
        fd = BytesIO()
        try:
            book = xlwt.Workbook()
            book.add_sheet(sheet)
            book.save(fd)
            fd.seek(0)
            self.open(fd)
        finally:
            fd.close()

    def open(self, path_or_file=None):
        path_or_file = path_or_file or self.path

        if hasattr(path_or_file, "read"):
            self._book = xlrd.open_workbook(
                file_contents=path_or_file.read(), on_demand=True, formatting_info=True,
            )
        else:
            self._book = xlrd.open_workbook(
                filename=path_or_file, on_demand=True, formatting_info=True
            )

    def close(self):
        self._book.release_resources()
        self._book = None
        self._active = None

    @contextmanager
    def _book_write(self):
        book = xlutils_copy(self._book)
        yield book

        fd = BytesIO()
        try:
            book.save(fd)
            fd.seek(0)
            self.close()
            self.open(fd)
        finally:
            fd.close()

    def save(self, path=None):
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")

        book = xlutils_copy(self._book)
        book.save(path)

    def create_worksheet(self, name):
        with self._book_write() as book:
            book.add_sheet(name)

        self.active = name

    def read_worksheet(self, name=None, header=False):
        name = self._get_sheetname(name)
        sheet = self._book.sheet_by_name(name)
        start, columns = self._get_columns(sheet, header)

        data = []
        for r in range(start, sheet.nrows):
            row = {}
            for c in range(sheet.ncols):
                column = columns[c]
                cell = sheet.cell(r, c)
                row[column] = self._parse_type(cell)
            data.append(row)

        self.active = name
        return data

    def _parse_type(self, cell):
        value = cell.value

        if cell.ctype == xlrd.XL_CELL_DATE:
            value = xlrd.xldate_as_datetime(value, self._book.datemode)
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            value = bool(value)
        elif cell.ctype == xlrd.XL_CELL_ERROR:
            value = xlrd.biffh.error_text_from_code.get(value, "#ERROR")
        elif cell.ctype == xlrd.XL_CELL_NUMBER and value.is_integer():
            value = int(value)

        return value

    def append_worksheet(self, name=None, content=None, header=False):
        content = content or []
        name = self._get_sheetname(name)
        sheet_read = self._book.sheet_by_name(name)
        _, columns = self._get_columns(sheet_read, header)

        with self._book_write() as book:
            sheet_write = book.get_sheet(name)

            for r, row in enumerate(content, sheet_read.nrows):
                if isinstance(row, (list, tuple)):
                    for c, value in enumerate(row):
                        sheet_write.write(r, c, value)
                elif isinstance(row, dict):
                    for column, value in row.items():
                        sheet_write.write(r, columns.index[column], value)
                else:
                    raise ValueError(f"Unknown row type: {type(row)}")

        self.active = name

    def remove_worksheet(self, name=None):
        name = self._get_sheetname(name)

        if name == self.active:
            self.active = self.sheetnames[0]

        with self._book_write() as book:
            # This is pretty ugly, but there seems to be no other way to
            # remove sheets from the xlwt.Workbook instance
            # pylint: disable=protected-access
            book._Workbook__worksheets = [
                sheet for sheet in book._Workbook__worksheets if sheet.name != name
            ]

    def rename_worksheet(self, title, name=None):
        title = str(title)
        name = self._get_sheetname(name)

        with self._book_write() as book:
            sheet = book.get_sheet(name)
            sheet.name = title

        self.active = title
