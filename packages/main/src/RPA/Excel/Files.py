# pylint: disable=too-many-lines
import logging
import pathlib
import re
from collections import defaultdict
from contextlib import contextmanager
from io import BytesIO
from typing import Any, List, Optional, Union

import openpyxl
import xlrd
import xlwt

from PIL import Image, ImageColor
from xlutils.copy import copy as xlutils_copy
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import cell as utils_cell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color as xlsColor
from openpyxl.formula.translate import Translator
from RPA.Tables import Table, Tables, return_table_as_raw_list


PathType = Union[str, pathlib.Path]


def get_column_index(column: str) -> int:
    """Get column index from name, e.g. A -> 1, D -> 4, AC -> 29.
    Reverse of `get_column_letter()`
    """
    column = str(column).lower()

    col = 0
    for digit, char in enumerate(column[::-1]):
        value = ord(char) - 96
        col += (26**digit) * value

    return col


def ensure_unique(values: Any) -> List[Any]:
    """Ensures that each string value in the list is unique.
    Adds a suffix to each value that has duplicates,
    e.g. [Banana, Apple, Lemon, Apple] -> [Banana, Apple, Lemon, Apple_2]
    """

    def to_unique(values: Any) -> List[Any]:
        output = []
        seen = defaultdict(int)
        for value in values:
            if seen[value] and isinstance(value, str):
                output.append("%s_%d" % (value, seen[value] + 1))
            else:
                output.append(value)
            seen[value] += 1
        return output

    # Repeat process until each column is unique
    output = to_unique(values)
    while True:
        verify = to_unique(output)
        if output == verify:
            break
        output = verify

    return output


class Files:
    """The `Excel.Files` library can be used to read and write Excel
    files without the need to start the actual Excel application.

    It supports both legacy ``.xls`` files and modern ``.xlsx`` files.

    **Note:** To run macros or load password protected worksheets,
    please use the Excel application library.

    **Examples**

    **Robot Framework**

    A common use-case is to load an existing Excel file as a table,
    which can be iterated over later in a Robot Framework keyword or task:

    .. code-block:: robotframework

        *** Settings ***
        Library    RPA.Tables
        Library    RPA.Excel.Files

        *** Keywords ***
        Read orders as table
            Open workbook    ${ORDERS_FILE}
            ${worksheet}=    Read worksheet   header=${TRUE}
            ${orders}=       Create table     ${worksheet}
            [Return]         ${orders}
            [Teardown]       Close workbook

    Processing all worksheets in the Excel file and checking row count:

    .. code-block:: robotframework

        *** Settings ***
        Library    RPA.Excel.Files

        *** Variables ***
        ${EXCEL_FILE}   /path/to/excel.xlsx

        *** Tasks ***
        Rows in the sheet
            [Setup]      Open Workbook    ${EXCEL_FILE}
            @{sheets}=   List Worksheets
            FOR  ${sheet}  IN   @{sheets}
                ${count}=  Get row count in the sheet   ${sheet}
                Log   Worksheet '${sheet}' has ${count} rows
            END

        *** Keywords ***
        Get row count in the sheet
            [Arguments]      ${SHEET_NAME}
            ${sheet}=        Read Worksheet   ${SHEET_NAME}
            ${rows}=         Get Length  ${sheet}
            [Return]         ${rows}

    Creating a new Excel file with a dictionary:

    .. code-block:: robotframework

        *** Tasks ***
        Creating new Excel
            Create Workbook  my_new_excel.xlsx
            FOR    ${index}    IN RANGE    20
                &{row}=       Create Dictionary
                ...           Row No   ${index}
                ...           Amount   ${index * 25}
                Append Rows to Worksheet  ${row}  header=${TRUE}
            END
            Save Workbook

    Creating a new Excel file with a list:

    .. code-block:: robotframework

        *** Variables ***
        @{heading}   Row No   Amount
        @{rows}      ${heading}

        *** Tasks ***
        Creating new Excel
            Create Workbook  my_new_excel.xlsx
            FOR    ${index}    IN RANGE   1  20
                @{row}=         Create List   ${index}   ${index * 25}
                Append To List  ${rows}  ${row}
            END
            Append Rows to Worksheet  ${rows}
            Save Workbook

    **Python**

    The library can also be imported directly into Python.

    .. code-block:: python

        from RPA.Excel.Files import Files

        def read_excel_worksheet(path, worksheet):
            lib = Files()
            lib.open_workbook(path)
            try:
                return lib.read_worksheet(worksheet)
            finally:
                lib.close_workbook()
    """

    ROBOT_LIBRARY_SCOPE = "GLOBAL"
    ROBOT_LIBRARY_DOC_FORMAT = "REST"

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.workbook = None

    def _require_open_xlsx_workbook(self, keyword_name: str):
        assert self.workbook, "No active workbook"
        if isinstance(self.workbook, XlsWorkbook):
            raise NotImplementedError(
                f"Keyword '{keyword_name}' does not support XLS files"
            )

    def _load_workbook(
        self, path: str, data_only: bool, read_only: bool
    ) -> Union["XlsWorkbook", "XlsxWorkbook"]:
        # pylint: disable=broad-except
        path = pathlib.Path(path).resolve(strict=True)

        try:
            book = XlsxWorkbook(path)
            book.open(data_only=data_only, read_only=read_only)
            return book
        except InvalidFileException as exc:
            self.logger.debug(exc)  # Unsupported extension, silently try xlrd
        except Exception as exc:
            self.logger.info(
                "Failed to open as Office Open XML (.xlsx) format: %s", exc
            )

        try:
            book = XlsWorkbook(path)
            book.open()
            return book
        except Exception as exc:
            self.logger.info("Failed to open as Excel Binary Format (.xls): %s", exc)

        raise ValueError(
            f"Failed to open Excel file ({path}), "
            "verify that the path and extension are correct"
        )

    def create_workbook(
        self,
        path: Optional[str] = None,
        fmt: str = "xlsx",
        sheet_name: Optional[str] = None,
    ) -> Union["XlsWorkbook", "XlsxWorkbook"]:
        """Create and open a new Excel workbook.

        Automatically also creates a new worksheet with the name `sheet_name`.
        (defaults to "Sheet")

        **Note:** Must be paired with the ``Save Workbook`` keyword
        or the newly created workbook will be deleted upon robot completion.

        **Note:** The file name/path must be set in either the ``Create Workbook``
        keyword or the ``Save Workbook`` keyword and must include the file extension.

        :param path: Save path for workbook; defaults to robot root if not provided.
        :param fmt: Format of workbook, i.e. xlsx or xls; Defaults to xlsx if not
            provided.
        :param sheet_name: Custom name for the initial sheet.
        :return: Workbook object.

        Examples:

        .. code-block:: robotframework

            # Create modern format workbook.
            Create Workbook
            Save Workbook    orders.xlsx

            # Create modern format workbook with custom sheet name.
            Create Workbook  sheet_name=MyCustomSheetName
            Save Workbook    orders.xlsx

            # Create modern format workbook with a path set.
            Create Workbook    path=${OUTPUT_DIR}${/}orders.xlsx
            Save Workbook

            # Create legacy format workbook.
            Create Workbook    fmt=xls
            Save Workbook    orders.xls

            # Create legacy format workbook with a path set.
            # Note that the file name must be set in the `Create Workbook` keyword
            #  if the path argument is used.
            Create Workbook    path=${OUTPUT_DIR}${/}orders.xls    fmt=xls
            Save Workbook

        .. code-block:: python

            # Create modern format workbook with defaults.
            lib = Files()
            lib.create_workbook()
            lib.save_workbook("orders.xlsx")

            # Create modern format workbook with a path set.
            lib = Files()
            lib.create_workbook(path="./output/orders.xlsx", fmt="xlsx")
            lib.save_workbook()

            # Create legacy format workbook.
            lib = Files()
            lib.create_workbook(fmt="xls")
            lib.save_workbook("orders.xls")

            # Create legacy format workbook with a path set.
            # Note that the file name must be set in the `Create Workbook` keyword
            #  if the path is used.
            lib = Files()
            lib.create_workbook(path="./output/orders.xls", fmt="xls")
            lib.save_workbook()
        """
        if self.workbook:
            self.close_workbook()

        fmt = str(fmt).lower().strip()
        if fmt == "xlsx":
            self.workbook = XlsxWorkbook(path)
        elif fmt == "xls":
            self.workbook = XlsWorkbook(path)
        else:
            raise ValueError(f"Unknown format: {fmt}")

        self.workbook.create()
        if sheet_name is not None:
            self.rename_worksheet(self.get_active_worksheet(), sheet_name)

        return self.workbook

    def open_workbook(
        self,
        path: str,
        data_only: Optional[bool] = False,
        read_only: Optional[bool] = False,
    ) -> Union["XlsWorkbook", "XlsxWorkbook"]:
        """Open an existing Excel workbook.

        Opens the workbook in memory and sets it as the active workbook.
        **This library can only have one workbook open at a time, and
        any previously opened workbooks are closed first.**

        The file can be in either ``.xlsx`` or ``.xls`` format.

        :param path: path to Excel file
        :param data_only: controls whether cells with formulas have either
         the formula (default, False) or the value stored the last time Excel
         read the sheet (True). Affects only ``.xlsx`` files.
        :return: Workbook object

        Examples:

        .. code-block:: robotframework

            # Open workbook with only path provided
            Open Workbook    path/to/file.xlsx

            # Open workbook with path provided and reading formulas in cells
            # as the value stored
            # Note: Can only be used with XLSX workbooks
            Open Workbook    path/to/file.xlsx    data_only=True

        .. code-block:: python

            # Open workbook with only path provided
            lib.open_workbook(path="path/to/file.xlsx")

            # Open workbook with path provided and reading formulas in cells
            # as the value stored
            # Note: Can only be used with XLSX workbooks
            lib.open_workbook(path="path/to/file.xlsx", data_only=True)
        """
        if self.workbook:
            self.close_workbook()

        self.workbook = self._load_workbook(path, data_only, read_only)
        self.logger.info("Opened workbook: %s", self.workbook)
        return self.workbook

    def close_workbook(self) -> None:
        """Close the active workbook.

        Examples:

        .. code-block:: robotframework

            # Close active workbook
            Close Workbook

        .. code-block:: python

            # Close active workbook
            lib.close_workbook()
        """

        if self.workbook:
            self.logger.info("Closing workbook: %s", self.workbook)
            self.workbook.close()
            self.workbook = None

    def save_workbook(
        self, path: Optional[str] = None
    ) -> Union["XlsWorkbook", "XlsxWorkbook"]:
        """Save the active workbook.

        **Note:** No changes to the workbook are saved to the actual file unless
        this keyword is called.

        :param path: Path to save to. If not given, uses path given
                     when opened or created.
        :return:     Workbook object

        Examples:

        .. code-block:: robotframework

            # Saving the active workbook to a new location/filename or saving to
            # a new location/filename
            # Note: You cannot use Save Workbook to convert from XLSX to XLS
            # or vice-versa
            Save Workbook   path=${OUTPUT_DIR}${/}orders.xlsx

            # Saving the active workbook changes if location/filename were set
            # in Create Workbook or Open Workbook
            Save Workbook

        .. code-block:: python

            # Saving the active workbook to a new location/filename or saving to
            # a new location/filename
            # Note: You cannot use Save Workbook to convert from XLSX to XLS
            # or vice-versa
            lib.save_workbook(path="./output/orders.xlsx")

            # Saving the active workbook changes if location/filename were set
            # in Create Workbook or Open Workbook
            lib.save_workbook()

        """
        assert self.workbook, "No active workbook"

        try:
            extension = pathlib.Path(path).suffix
        except TypeError:
            extension = None

        if (
            self.workbook.extension is not None
            and extension is not None
            and self.workbook.extension != extension
        ):
            self.logger.warning(
                "Changed file extension from %s to %s",
                self.workbook.extension,
                extension,
            )

        self.workbook.validate_content()
        return self.workbook.save(path)

    def list_worksheets(self) -> List[str]:
        """List all names of worksheets in the given workbook.

        :return:    List containing the names of the worksheets

        Examples:

        .. code-block:: robotframework

            # List Worksheets will read the worksheet names into a list variable
            # The variable should be declared with the List type "@" when being used
            # to store the sheet names from the List Worksets keyword
            @{sheets}=    List Worksheets

        .. code-block:: python

            # List Worksheets will read the worksheet names into a list variable
            # The variable should be declared with the List type "@" when being used
            # to store the sheet names from the List Worksets keyword
            sheets = lib.list_worksheets()

        """
        assert self.workbook, "No active workbook"
        return self.workbook.sheetnames

    def worksheet_exists(self, name: str) -> bool:
        """Return True if worksheet with given name is in workbook.

        :param name: Name of worksheet you are looking for
        :return: `True` if the worksheet exists, `False` otherwise

        Examples:

        .. code-block:: robotframework

            # To use Worksheet Exists in a conditional statement set it to
            # a variable first, like you see here, and then compare the
            # variable to ${TRUE} or ${FALSE}
            ${Does_Worksheet_Exist}=    Worksheet Exists    Sheet

        .. code-block:: python

            Does_Worksheet_Exist = lib.worksheet_exists("Sheet")
        """
        assert self.workbook, "No active workbook"
        return bool(str(name) in self.list_worksheets())

    def get_active_worksheet(self) -> str:
        """Get the name of the worksheet which is currently active.


        :return:    Active worksheet name

        Examples:

        .. code-block:: robotframework

            ${Active_Worksheet}=    Get Active Worksheet

        .. code-block:: python

            Active_Worksheet = lib.get_active_worksheet()
        """

        assert self.workbook, "No active workbook"
        return self.workbook.active

    def set_active_worksheet(self, value: Union[str, int]) -> None:
        """Set the active worksheet.

        This keyword can be used to set the default worksheet for keywords,
        which removes the need to specify the worksheet name for each keyword.
        It can always be overridden on a per-keyword basis.

        :param value: Index or name of worksheet

        Examples:

        .. code-block:: robotframework

            # Set using the name of the worksheet
            Set Active Worksheet    Customers

            # Set using the index of the worksheet
            # Worksheet index begings at 0
            Set Active Worksheet    2

        .. code-block:: python

            # Set using the name of the worksheet
            lib.set_active_worksheet("Customers")

            # Set using the index of the worksheet
            # Worksheet index begings at 0
            lib.set_active_worksheet(2)
        """
        assert self.workbook, "No active workbook"
        self.workbook.active = value

    def create_worksheet(
        self,
        name: str,
        content: Optional[Any] = None,
        exist_ok: Optional[bool] = False,
        header: Optional[bool] = False,
    ) -> None:
        """Create a new worksheet in the current workbook.

        :param name:     Name of new worksheet
        :param content:  Optional content for worksheet
        :param exist_ok: If `False`, raise an error if name is already in use
        :param header:   If content is provided, write headers to worksheet

        Examples:

        .. code-block:: robotframework

            # Create a new blank worksheet named "Customers"
            Create Worksheet    Customers

            # Create a new worksheet with headers and contents using
            # a List of Dictonaries
            # Don't forget to `Save Workbook` once your changes are complete
            &{Employees_Row1}=    Create Dictionary    name=Mark    age=${58}
            &{Employees_Row2}=    Create Dictionary    name=John    age=${22}
            &{Employees_Row3}=    Create Dictionary    name=Adam    age=${67}
            @{Worksheet_Data}=    Create List
            ...    ${Worksheet_Data_row1}
            ...    ${Worksheet_Data_row2}
            ...    ${Worksheet_Data_row3}
            Create Worksheet
            ...    name=Employees
            ...    content=${Worksheet_Data}
            ...    header=True
            Save Workbook

            # Create a new workseet using a Dictionary of Lists
            # Don't forget to `Save Workbook` once your changes are complete
            @{Employees_name}=    Create List    Mark    John    Adam
            @{Employees_age}=    Create List    ${58}    ${22}    ${67}
            &{Worksheet_Data}=    Create Dictionary
            ...    name=${Worksheet_Data_name}
            ...    age=${Worksheet_Data_age}
            Create Worksheet
            ...    name=Employees
            ...    content=${Worksheet_Data}
            ...    header=True
            Save Workbook

        .. code-block:: python

            # Create a new blank worksheet named "Customers"
            lib.create_worksheet("Customers")

            # Create a new workseet using a List of Dictionaries
            # Don't forget to `Save Workbook` once your changes are complete
            Worksheet_Data = [
                {"name": "Mark", "age": 58},
                {"name": "John", "age": 22},
                {"name": "Adam", "age": 67},
                ]
            lib.create_worksheet(name="Employees",content=Worksheet_Data,header=True)
            lib.save_workbook()

            # Create a new workseet using a Dictionary of Lists
            # Don't forget to `Save Workbook` once your changes are complete
            Worksheet_Data = {
                "name": ["Mark", "John", "Adam"],
                "age":  [    58,     22,     67],
                }
            lib.create_worksheet(name="Employees",content=Worksheet_Data,header=True)
            lib.save_workbook()
        """
        assert self.workbook, "No active workbook"
        if name in self.workbook.sheetnames and not exist_ok:
            raise ValueError(f"Sheet with name {name} already exists")

        self.workbook.create_worksheet(name)
        if content:
            self.workbook.append_worksheet(name, content, header)

    def read_worksheet(
        self,
        name: Optional[str] = None,
        header: Optional[bool] = False,
        start: Optional[int] = None,
    ) -> List[dict]:
        """Read the content of a worksheet into a list of dictionaries.

        Each key in the dictionary will be either values from the header row,
        or Excel-style column letters.

        :param name:   Name of worksheet to read (optional).
                       Defaults to the active worksheet.
        :param header: If `True`, use the first row of the worksheet
                       as headers for the rest of the rows. Default is `False`.
        :param start:  Row index to start reading data from (1-indexed).
                       Default value is row 1.
        :return:       List of dictionaries that represents the worksheet

        Examples:

        .. code-block:: robotframework

            # The most simple form. Column keys will be Column letters.
            ${rows}=    Read Worksheet

            # Since `header=True` the keys will be the header values
            ${rows}=    Read Worksheet     header=True

            # Uses the header values as keys and starts reading at row 3
            ${rows}=    Read Worksheet     header=True    start=${3}

        .. code-block:: python

            # The most simple form. Keys will be Column letters.
            rows = lib.read_worksheet()

            # Since `header=True` the keys will be the header values
            rows = lib.read_worksheet(header=True)

            # Uses the header values as keys and starts reading at row 3
            rows = lib.read_worksheet(header=True, start=3)
        """
        assert self.workbook, "No active workbook"
        return self.workbook.read_worksheet(name, header, start)

    def read_worksheet_as_table(
        self,
        name: Optional[str] = None,
        header: bool = False,
        trim: bool = True,
        start: Optional[int] = None,
    ) -> Table:
        """Read the contents of a worksheet into a Table container. Allows
        sorting/filtering/manipulating using the ``RPA.Tables`` library.

        :param name:   Name of worksheet to read (optional).
                       Defaults to the active worksheet.
        :param header: If `True`, use the first row of the worksheet
                       as headers for the rest of the rows. Default value is False.
        :param trim:   Remove all empty rows from the end of the worksheet.
                       Default value is True.
        :param start:  Row index to start reading data from (1-indexed).
                       Default value is row 1.
        :return:       Table object that represents the worksheet

        Examples:

        .. code-block:: robotframework

            # The most simple form. Column keys will be Column letters.
            ${table}=    Read Worksheet As Table

            # Since `header=True` the keys will be the header values
            ${table}=    Read Worksheet As Table     header=True

            # Uses the header values as keys and starts reading at row 3
            ${table}=    Read Worksheet As Table     header=True    start=${3}

        .. code-block:: python

            # The most simple form. Keys will be Column letters.
            table = lib.read_worksheet_as_table()

            # Since `header=True` the keys will be the header values
            table = lib.read_worksheet_as_table(header=True)

            # Uses the header values as keys and starts reading at row 3
            table = lib.read_worksheet_as_table(header=True, start=3)
        """
        tables = Tables()
        sheet = self.read_worksheet(name, header, start)
        return tables.create_table(sheet, trim)

    def append_rows_to_worksheet(
        self,
        content: Any,
        name: Optional[str] = None,
        header: bool = False,
        start: Optional[int] = None,
        formatting_as_empty: Optional[bool] = False,
    ) -> List[dict]:
        """Append values to the end of the worksheet.

        :param content: Rows of values to append
        :param name:    Name of worksheet to append to (optional).
                        Defaults to the active worksheet.
        :param header:  Set rows according to existing header row
        :param start:   Start of data, NOTE: Only required when header is True
        :param formatting_as_empty: if True, the cells containing only
                        formatting (no values) are considered empty.
        :return:        List of dictionaries that represents the worksheet

        The ``content`` argument can be of any tabular format. Typically,
        this is a Table object created by the ``RPA.Tables`` library,
        but it can also be a list of lists, or a list of dictionaries.

        If the ``header`` flag is enabled, the existing header in the worksheet
        is used to insert values in the correct columns. This assumes that
        that source data has this data available.

        If the header is not on the first row of the worksheet,
        the ``start`` argument can be used to give the correct row index.

        Examples:

        .. code-block:: robotframework

            # Append an existing Table object
             # Create a new table using a Dictionary of Lists
            @{table_name}=    Create List    Sara    Beth    Amy
            @{table_age}=    Create List    ${48}    ${21}    ${57}
            &{table}=    Create Dictionary    name=${table_name}    age=${table_age}
            Create Table    ${table}
            Append rows to worksheet    ${table}
            Save Workbook

            # Append to a worksheet with headers on row 5
             # Create a new table using a Dictionary of Lists
            @{table_name}=    Create List    Sara    Beth    Amy
            @{table_age}=    Create List    ${48}    ${21}    ${57}
            &{table}=    Create Dictionary    name=${table_name}    age=${table_age}
            Create Table    ${table}
            Append rows to worksheet    ${table}    header=${TRUE}   start=5
            Save Workbook

        .. code-block:: python

            # Append an existing Table object
            table = {
                "name": ["Sara", "Beth", "Amy"],
                "age":  [    48,     21,     57],
                }
            lib.append_rows_to_worksheet(table)
            lib.save_workbook()

            # Append to a worksheet with headers on row 5
            table = {
                "name": ["Sara", "Beth", "Amy"],
                "age":  [    48,     21,     57],
                }
            lib.append_rows_to_worksheet(table, header=True, start=5)
            lib.save_workbook()
        """
        assert self.workbook, "No active workbook"
        return self.workbook.append_worksheet(
            name, content, header, start, formatting_as_empty
        )

    def remove_worksheet(self, name: str = None) -> None:
        """Remove a worksheet from the active workbook.

        :param name: Name of worksheet to remove (optional).
                     Defaults to the active worksheet.

        Examples:

        .. code-block:: robotframework

            # Remove last worksheet
            ${sheets}=       List worksheets
            Remove worksheet    ${sheets}[-1]

            # Remove worksheet by name
            Remove Worksheet    Sheet

        .. code-block:: python

            # Remove last worksheet
            sheets = lib.list_worksheets()
            lib.remove_worksheet(sheets[-1])

            # Remove worksheet by name
            lib.remove_worksheet("Sheet")
        """
        assert self.workbook, "No active workbook"
        self.workbook.remove_worksheet(name)

    def rename_worksheet(self, src_name: str, dst_name: str) -> None:
        """Rename an existing worksheet in the active workbook.

        :param src_name: Current name of worksheet
        :param dst_name: Future name of worksheet

        Examples:

        .. code-block:: robotframework

            Rename worksheet    Sheet    Orders

        .. code-block:: python

            lib.rename_worksheet("Sheet","Orders")
        """
        assert self.workbook, "No active workbook"
        self.workbook.rename_worksheet(dst_name, src_name)

    def find_empty_row(self, name: Optional[str] = None) -> int:
        """Find the first empty row after existing content,
        and return the row number.

        :param name:    Name of worksheet (optional). Defaults to the active worksheet.
        :return:        First row number of empty row

        Examples:

        .. code-block:: robotframework

            ${next}=    Find empty row

        .. code-block:: python

            next = lib.find_empty_row()
        """
        assert self.workbook, "No active workbook"
        return self.workbook.find_empty_row(name)

    def get_cell_value(
        self, row: int, column: Union[str, int], name: Optional[str] = None
    ) -> Any:
        """Get a cell value in the given worksheet.

        :param row:     Index of row to read, e.g. 3
        :param column:  Name or index of column, e.g. C or 7
        :param name:    Name of worksheet (optional). Defaults to active worksheet.
        :return:        Cell value

        Examples:

        .. code-block:: robotframework

            # Read header names
            ${column1}=    Get cell value    1    A
            ${column2}=    Get cell value    1    B
            ${column3}=    Get cell value    1    C

        .. code-block:: python

            # Read header names
            column1 = lib.get_cell_value(1, "A")
            column2 = lib.get_cell_value(1, "B")
            column3 = lib.get_cell_value(1, "C")
        """
        assert self.workbook, "No active workbook"
        return self.workbook.get_cell_value(row, column, name)

    def set_cell_value(
        self,
        row: int,
        column: Union[str, int],
        value: Any,
        name: Optional[str] = None,
        fmt: Optional[Union[str, float]] = None,
    ) -> None:
        """Set a cell value in the given worksheet.

        :param row:     Index of row to write, e.g. 3
        :param column:  Name or index of column, e.g. C or 7
        :param value:   New value of cell
        :param name:    Name of worksheet (optional). Defaults to active worksheet.
        :param fmt:     Format code for cell (optional)

        Examples:

        .. code-block:: robotframework

            # Set a value in the first row and column
            Set cell value    1    1    Some value
            Set cell value    1    A    Some value

            # Set a value with cell formatting
            Set cell value    2    B    ${value}    fmt=0%

        .. code-block:: python

            # Set a value in the first row and column
            lib.set_cell_value(1, 1, "Some value")
            lib.set_cell_value(1, "A", "Some value")

            # Set a value with cell formatting
            lib.set_cell_value(2, "B", value, fmt="0%")
        """
        assert self.workbook, "No active workbook"

        self.workbook.set_cell_value(row, column, value, name)

        if fmt is not None:
            self.workbook.set_cell_format(row, column, fmt, name)

    def set_cell_format(
        self,
        row: int,
        column: Union[str, int],
        fmt: Union[str, float],
        name: Optional[str] = None,
    ) -> None:
        """Set format for cell.

        Does not affect the values themselves, but changes how the values
        are displayed when opening with an external application such as
        Microsoft Excel or LibreOffice Calc.

        :param row:     Index of row to write, e.g. 3
        :param column:  Name or index of column, e.g. C or 7
        :param fmt:     Format code for cell
        :param name:    Name of worksheet (optional). Defaults to active worksheet.

        The ``fmt`` argument accepts all format code values that
        are supported by the aforementioned applications.

        Some examples of valid values:

        ======== ===========
        Format   Explanation
        ======== ===========
        0.00     Number with two decimal precision
        0%       Percentage without decimals
        MM/DD/YY Date with month, day, and year
        @        Text value
        BOOLEAN  Boolean value
        ======== ===========

        Examples:

        .. code-block:: robotframework

            # Set value to have one decimal precision
            Set cell format   2  B    00.0

        .. code-block:: python

            # Set value to have one decimal precision
            lib.set_cell_format(2, "B", 00.0)
        """
        assert self.workbook, "No active workbook"
        self.workbook.set_cell_format(row, column, fmt, name)

    def insert_image_to_worksheet(
        self,
        row: int,
        column: Union[int, str],
        path: str,
        scale: float = 1.0,
        name: Optional[str] = None,
    ) -> None:
        """Insert an image into the given cell.

        The ``path`` argument should be a local file path to the image file.

        By default the image is inserted in the original size, but it can
        be scaled with the ``scale`` argument. It's scaled with a factor
        where the value ``1.0`` is the default.

        :param row:     Index of row to write
        :param column:  Name or index of column
        :param path:    Path to image file
        :param scale:   Scale of image (optional). Default value is "1.0".
        :param name:    Name of worksheet (optional). Defaults to the active worksheet.

        Examples:

        .. code-block:: robotframework

            Insert image to worksheet    ${last_row}    A    screenshot.png

        .. code-block:: python

            lib.insert_image_to_worksheet(last_row, "A", "screenshot.png")
        """
        assert self.workbook, "No active workbook"
        image = Image.open(path)

        if scale != 1.0:
            fmt = image.format
            width = int(image.width * float(scale))
            height = int(image.height * float(scale))
            image = image.resize((width, height), Image.Resampling.LANCZOS)
            image.format = fmt

        self.workbook.insert_image(row, column, image, name)

    def get_worksheet_value(
        self, row: int, column: Union[str, int], name: Optional[str] = None
    ) -> Any:
        """Alias for keyword ``Get cell value``, see the original keyword
        for documentation.
        """
        # Old keyword name, deprecate at some point
        return self.get_cell_value(row, column, name)

    def set_worksheet_value(
        self,
        row: int,
        column: Union[str, int],
        value: Any,
        name: Optional[str] = None,
        fmt: Optional[Union[str, float]] = None,
    ) -> Any:
        """Alias for keyword ``Set cell value``, see the original keyword
        for documentation.
        """
        # Old keyword name, deprecate at some point
        return self.set_cell_value(row, column, value, name, fmt)

    def clear_cell_range(
        self,
        range_string: str,
    ):
        """Clear cell values for a given range.

        :param range_string: single cell or range of cells

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            # area of cells
            Clear Cell Range    A9:A100
            # single cell
            Clear Cell Range    A2

        Python example.

        .. code-block:: python

            lib.clear_cell_range("A1")
            lib.clear_cell_range("B2:B50")
        """
        self._require_open_xlsx_workbook("clear_cell_range")
        cr = CellRange(range_string=range_string)
        for row, acell in list(cr.cells):
            self.workbook.book.active.cell(row, acell).value = None

    def delete_rows(self, start: int, end: Optional[int] = None):
        """Delete row or rows beginning from start row number to
        possible end row number.

        :param start: row number to start deletion from
        :param end: optional row number for last row to delete

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Delete Rows   2       # delete row 2
            Delete Rows   5  10   # delete rows 5-10

        Python example.

        .. code-block:: python

            lib.delete_rows(2)
            lib.delete_rows(5,10)
        """
        self._require_open_xlsx_workbook("delete_rows")
        amount = (end - start + 1) if end else 1
        self.workbook.book.active.delete_rows(start, amount)

    def delete_columns(
        self, start: Union[int, str], end: Optional[Union[int, str]] = None
    ):
        """Delete column or columns beginning from start column number/name to
        possible end column number/name.

        :param start: column number or name to start deletion from
        :param end: optional column number or name for last column to delete

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Delete Columns   C       # delete column C
            Delete Columns   3       # delete column 3 (same as C)
            Delete Columns   E  AA   # delete rows E-AA

        Python example.

        .. code-block:: python

            lib.delete_columns("D")
            lib.delete_rows(1, "JJ")
        """
        self._require_open_xlsx_workbook("delete_columns")
        start_column_index = (
            start if isinstance(start, int) else column_index_from_string(start)
        )
        amount = 1
        if end:
            end_column_index = (
                end if isinstance(end, int) else column_index_from_string(end)
            )
            amount = end_column_index - start_column_index + 1
        self.workbook.book.active.delete_cols(start_column_index, amount)

    def insert_columns_before(self, column: Union[int, str], amount: int = 1):
        """Insert column or columns before a column number/name.

        :param column: insert before this column
        :param amount: number of columns to insert, default 1

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Insert Columns Before   C      # insert 1 column before column C
            Insert Columns Before   A  3   # insert 3 columns before column A

        Python example.

        .. code-block:: python

            lib.insert_columns_before("C")
            lib.insert_columns_before("A", 3)
        """
        self._require_open_xlsx_workbook("insert_columns_before")

        column_index = (
            column_index_from_string(column) if isinstance(column, str) else column
        )
        self.workbook.book.active.insert_cols(column_index, amount)

    def insert_columns_after(self, column: Union[int, str], amount: int = 1):
        """Insert column or columns after a column number/name.

        :param column: insert after this column
        :param amount: number of columns to insert, default 1

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Insert Columns After   C      # insert 1 column after column C
            Insert Columns Before   A  3   # insert 3 columns after column A

        Python example.

        .. code-block:: python

            lib.insert_columns_after("C")
            lib.insert_columns_after("A", 3)
        """
        self._require_open_xlsx_workbook("insert_columns_after")

        column_index = (
            column_index_from_string(column) if isinstance(column, str) else column
        )
        self.workbook.book.active.insert_cols(column_index + amount - 1, amount)

    def insert_rows_before(self, row: int, amount: int = 1):
        """Insert row or rows before a row number.

        :param row: insert before this row
        :param amount: number of rows to insert, default 1

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Insert Rows Before   3      # insert 1 row before row 3
            Insert Rows Before   1  3   # insert 3 rows before row 1

        Python example.

        .. code-block:: python

            lib.insert_rows_before(1)
            lib.insert_rows_before(1, 3)
        """
        self._require_open_xlsx_workbook("insert_rows_before")

        self.workbook.book.active.insert_rows(row, amount)

    def insert_rows_after(self, row: int, amount: int = 1):
        """Insert row or rows after a row number.

        :param row: insert after this row
        :param amount: number of rows to insert, default 1

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Insert Rows After   3      # insert 1 row after row 3
            Insert Rows After   1  3   # insert 3 rows after row 1

        Python example.

        .. code-block:: python

            lib.insert_rows_after(1)
            lib.insert_rows_after(1, 3)
        """
        self._require_open_xlsx_workbook("insert_rows_after")

        self.workbook.book.active.insert_rows(row + amount - 1, amount)

    def copy_cell_values(self, source_range: str, target: str):
        """Copy cells from source to target.

        :param source_range: single cell or range of cells
        :param target: copy to this cell

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Copy Cell Values   A1:D4   G10

        Python example.

        .. code-block:: python

            lib.copy_cell_values("A1:D4", "G10")
        """
        self._require_open_xlsx_workbook("copy_cell_values")

        cr = CellRange(range_string=source_range)
        cells = list(cr.cells)
        target_cell_unpacked = utils_cell.coordinate_from_string(target)
        target_column = column_index_from_string(target_cell_unpacked[0])
        target_row = target_cell_unpacked[1]

        last_row = None
        column_index = 0
        for row, column in cells:
            if not last_row:
                last_row = row
            if last_row != row:
                target_row += 1
                last_row = row
                column_index = 0
            target_cell = self.workbook.book.active.cell(
                target_row, target_column + column_index
            )
            source_cell = self.workbook.book.active.cell(row, column)
            target_cell.value = source_cell.value
            column_index += 1

    def set_styles(
        self,
        range_string: str,
        font_name: str = None,
        family: str = None,
        size: int = None,
        bold: bool = False,
        italic: bool = False,
        underline: bool = False,
        strikethrough: bool = False,
        cell_fill: str = None,
        color: str = None,
        align_horizontal: str = None,
        align_vertical: str = None,
        number_format: str = None,
    ):
        """Set styles for range of cells.

        Possible values for the `align_horizontal`:

            - general
            - left
            - center
            - right
            - fill
            - justify
            - centerContinuous
            - distributed

        Possible values for the `align_vertical`:

            - top
            - center
            - bottom
            - justify
            - distributed

        Some examples for `number_formats`:

            - General
            - 0
            - 0.00
            - #,##0
            - #,##0.00
            - "$"#,##0_);("$"#,##0)
            - "$"#,##0_);[Red]("$"#,##0)
            - 0%
            - 0.00%
            - 0.00E+00
            - # ?/?
            - # ??/??
            - mm-dd-yy
            - d-mmm-yy
            - d-mmm
            - h:mm AM/PM
            - h:mm:ss AM/PM
            - h:mm
            - h:mm:ss
            - m/d/yy h:mm

        :param range_string: single cell or range of cells
        :param font_name: name of the font
        :param family: font family name
        :param size: size for the font
        :param bold: font style bold
        :param italic: font style italics
        :param underline: font style underline
        :param strikethrough: font style strikethrough
        :param cell_fill: cell fill color, in hex or color name
        :param color: font color, in hex or color name
        :param align_horizontal: cell horizontal alignment
        :param align_vertical: cell vertical alignment
        :param number_format: cell number format

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Set Styles    A1:D4
            ...  bold=True
            ...  cell_fill=lightblue
            ...  align_horizontal=center
            ...  number_format=h:mm AM/PM

            Set Styles    E2
            ...  strikethrough=True
            ...  color=FF0000

        Python example.

        .. code-block:: python

            lib.set_styles("A1:D4", bold=True, font_name="Arial", size=24)
        """
        self._require_open_xlsx_workbook("set_styles")

        cr = CellRange(range_string=range_string)
        font_parameters = {}
        self._set_font_param_if_given(font_parameters, "name", font_name)
        self._set_font_param_if_given(font_parameters, "family", family)
        self._set_font_param_if_given(font_parameters, "sz", size)
        self._set_font_param_if_given(font_parameters, "b", bold)
        self._set_font_param_if_given(font_parameters, "i", italic)
        self._set_font_param_if_given(font_parameters, "u", underline)
        self._set_font_param_if_given(font_parameters, "strike", strikethrough)
        self._set_color_if_given(font_parameters, color)

        for row, column in list(cr.cells):
            active_cell = self.workbook.book.active.cell(row, column)
            active_cell.font = Font(**font_parameters)
            self._set_fill_color(active_cell, cell_fill)
            self._set_cell_alignments(active_cell, align_horizontal, align_vertical)
            self._set_cell_number_format(active_cell, number_format)

    def _set_font_param_if_given(self, parameters, param_name, value):
        if value:
            parameters[param_name] = value

    def _set_color_if_given(self, parameters, value):
        if value:
            match = re.search(r"^(?:[0-9a-fA-F]{3}){2}$", value)
            if match:
                color_hex = value
            else:
                color = ImageColor.getrgb(value)
                color_hex = "%02x%02x%02x" % color
            parameters["color"] = xlsColor(color_hex)

    def _set_fill_color(self, active_cell, value):
        if value:
            match = re.search(r"^(?:[0-9a-fA-F]{3}){2}$", value)
            if match:
                color_hex = value
            else:
                color = ImageColor.getrgb(value)
                color_hex = "%02x%02x%02x" % color
            active_cell.fill = PatternFill("solid", color_hex, color_hex)

    def _set_cell_alignments(self, active_cell, horizontal, vertical):
        if horizontal or vertical:
            active_cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    def _set_cell_number_format(self, active_cell, number_format):
        if number_format:
            active_cell.number_format = number_format

    def auto_size_columns(
        self,
        start_column: Union[int, str],
        end_column: Optional[Union[int, str]] = None,
        width: Optional[int] = None,
    ):
        """Auto size column widths.

        Note. non-default font sizes might cause auto sizing issues

        :param start_column: column number or name to start from
        :param end_column: optional column number or name for last column
        :param width: if given will resize columns to this size, otherwise
         will auto_size

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Auto Size Columns   A   D    # will try auto size
            Auto Size Columns   B   D   16  # will set A-D columns sizes to 16
            Auto Size Columns   A   width=24  # will set column A size to 24

        Python example.

        .. code-block:: python

            lib.auto_size_columns("A", "D")
            lib.auto_size_columns("C", width=40)
        """
        self._require_open_xlsx_workbook("auto_size_columns")
        start_index = (
            column_index_from_string(start_column)
            if isinstance(start_column, str)
            else start_column
        )
        end_index = start_index
        if end_column:
            end_index = (
                column_index_from_string(end_column)
                if isinstance(end_column, str)
                else end_column
            )

        for col in range(start_index, end_index + 1):
            col_letter = get_column_letter(col)
            if width:
                self.workbook.book.active.column_dimensions[col_letter].width = width
            else:
                self.workbook.book.active.column_dimensions[col_letter].auto_size = True

    def hide_columns(
        self,
        start_column: Union[int, str],
        end_column: Optional[Union[int, str]] = None,
    ):
        """Hide column or columns in worksheet.

        :param start_column: column number or name to start from
        :param end_column: optional column number or name for last column

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Hide Columns   A   D    # hide columns A-D
            Hide Columns   A        # hide column A

        Python example.

        .. code-block:: python

            lib.hide_columns("A", "D")
            lib.hide_columns("A")
        """
        self._require_open_xlsx_workbook("hide_columns")

        self._set_column_hidden(True, start_column, end_column)

    def unhide_columns(
        self,
        start_column: Union[int, str],
        end_column: Optional[Union[int, str]] = None,
    ):
        """Unhide column or columns in worksheet.

        :param start_column: column number or name to start from
        :param end_column: optional column number or name for last column

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            Unhide Columns   A   D    # unhide columns A-D
            Unhide Columns   A        # unhide column A

        Python example.

        .. code-block:: python

            lib.unhide_columns("A", "D")
            lib.unhide_columns("A")
        """
        self._require_open_xlsx_workbook("unhide_columns")

        self._set_column_hidden(False, start_column, end_column)

    def _set_column_hidden(self, hidden: bool, start_column, end_column):
        start_index = (
            column_index_from_string(start_column)
            if isinstance(start_column, str)
            else start_column
        )
        end_index = start_index
        if end_column:
            end_index = (
                column_index_from_string(end_column)
                if isinstance(end_column, str)
                else end_column
            )

        for col in range(start_index, end_index + 1):
            col_letter = get_column_letter(col)
            self.workbook.book.active.column_dimensions[col_letter].hidden = hidden

    def set_cell_formula(
        self, range_string: str, formula: str, transpose: bool = False
    ):
        """Set cell formula for given range of cells.

        If `transpose` is set then formula is set for first cell of the
        range and the rest of cells will transpose the function to match
        to that cell.

        Otherwise (by default) all cells will get the same formula.

        :param range_string: cell range
        :param formula: formula for the cell
        :param transpose: on True the cell formulas will be transposed

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            # all cells will have same formula
            Set Cell Formula   E2:E10    =B2+5
            # cells will have transposed formulas
            # E2 will have =B2+5
            # E3 will have =B3+5
            # etc
            Set Cell Formula   E2:E10    =B2+5   True

        Python example.

        .. code-block:: python

            lib.set_cell_formula("E2:E10", "=B2+5")
            lib.set_cell_formula("E2:E10", "=B2+5", True)
        """
        self._require_open_xlsx_workbook("set_cell_formula")

        cr = CellRange(range_string=range_string)
        start_col, start_row, _, _ = cr.bounds
        start_col_str = f"{get_column_letter(start_col)}{start_row}"
        cells = list(cr.cells)

        for index, acell in enumerate(cells):
            row, column = acell
            col_str = f"{get_column_letter(column)}{row}"
            if (transpose and index == 0) or not transpose:
                self.workbook.book.active[col_str].value = formula
            elif transpose and index > 0:
                self.workbook.book.active[col_str] = Translator(
                    formula, origin=start_col_str
                ).translate_formula(col_str)

    def move_range(
        self, range_string: str, rows: int = 0, columns: int = 0, translate: bool = True
    ):
        """Move range of cells by given amount of rows and columns.

        Formulas are translated to match new location by default.

        *Note*. There is a bug in the openpyxl on moving negative rows/columns.

        :param range_string: cell range
        :param rows: number of rows to move
        :param columns: number of columns to move
        :param translate: are formulas translated for a new location

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            # move range 4 rows down
            Move Range   E2:E10    rows=4
            # move range 2 rows down, 2 columns right
            Move Range   E2:E10    rows=2  columns=2

        Python example.

        .. code-block:: python

            lib.move_range("E2:E10", rows=4)
            lib.move_range("E2:E10", rows=2, columns=2)
        """
        self._require_open_xlsx_workbook("move_range")

        self.workbook.book.active.move_range(
            range_string, rows=rows, cols=columns, translate=translate
        )

    def set_cell_values(
        self, start_cell: str, values: Union[list, Table], table_heading: bool = False
    ):
        """Set cell values given as list of lists or as a `RPA.Tables.Table`.

        *Note.* Will overwrite cells if table structure causes cells to overlap.

        :param start_cell: starting cell in a string
        :param values: list of lists or a Table
        :param table_heading: if values are given as a Table, this parameter
         defines if Table headings should be inserted as a row

        Examples:

        Robot Framework example.

        .. code-block:: robotframework

            @{all_rows}=    Create List
            ${headers}=    Create List    first    second   third  fourth
            FOR    ${num}    IN RANGE    1    2000
                @{row}=    Create List    ${num}    ${num+1}    ${num*2}    ${num*4}
                Append To List    ${all_rows}    ${row}
            END
            #  Set Cell Values from Table (include headers)
            ${table}=    Create Table    ${all_rows}    columns=${headers}
            Set Cell Values   G1   ${table}   True
            #  Set Cell Values from a list of lists
            # uncomment if headings should be added
            # Append To List  ${all_rows}   ${headers}
            Set Cell Values   M1   ${all_rows}

            # Simplest form of adding values
            @{values}=    Evaluate    [[1,2,3],[4,5,6],['a','b','c','d']]
            Set Cell Values   A1   ${values}

        Python example.

        .. code-block:: python

            data =  [[1,2,3],[4,5,6],['a','b','c','d']]
            lib.set_cell_values("E2", data)
        """
        self._require_open_xlsx_workbook("set_cell_values")
        if isinstance(values, Table):
            data_values = return_table_as_raw_list(values, table_heading)
        else:
            data_values = values
        cr = CellRange(range_string=start_cell)
        start_col, start_row, _, _ = cr.bounds
        for row_index, row in enumerate(data_values):
            if isinstance(row, list):
                for col_index, col in enumerate(row):
                    column = start_col + col_index
                    row = start_row + row_index
                    self.workbook.book.active[
                        f"{get_column_letter(column)}{row}"
                    ].value = col
            else:
                column = start_col + row_index
                row = start_row
                self.workbook.book.active[
                    f"{get_column_letter(column)}{row}"
                ].value = row


class BaseWorkbook:

    """Common logic for both .xls and .xlsx files management."""

    def __init__(self, path: Optional[PathType] = None):
        self.logger = logging.getLogger(__name__)
        self.path = path
        self._book = None
        self._extension = None
        self._active = None

    @property
    def book(self):
        return self._book

    def _validate_content(self, props_obj: Any):
        # Strips leading/trailing whitespace in Excel properties.
        public_props = [prop for prop in dir(props_obj) if not prop.startswith("_")]
        for prop in public_props:
            value = getattr(props_obj, prop)
            if value and isinstance(value, str):
                setattr(props_obj, prop, value.strip())


class XlsxWorkbook(BaseWorkbook):
    """Container for manipulating modern Excel files (.xlsx)"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._read_only = False

    @staticmethod
    def is_sheet_empty(sheet):
        # Maximum rows/columns are always 1 or more, even when the sheet doesn't
        #  contain cells at all. (https://stackoverflow.com/a/37673211/4766178)
        # _cells does not exist for the ReadOnlyWorksheet, no way of knowing
        # is there data or not
        # pylint: disable=protected-access
        is_readonly = isinstance(sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet)
        return (
            False if is_readonly else not sheet._cells
        )  # there's no public API for this

    @property
    def sheetnames(self):
        return list(self._book.sheetnames)

    @property
    def active(self):
        if not self._active:
            self._active = self._book.active.title

        return self._active

    @active.setter
    def active(self, value):
        if isinstance(value, int):
            value = self.sheetnames[value]
        elif value not in self.sheetnames:
            raise ValueError(f"Unknown worksheet: {value}")

        self._book.active = self.sheetnames.index(value)
        self._active = value

    @property
    def extension(self):
        return self._extension

    @property
    def read_only(self):
        return self._read_only

    def _get_sheetname(self, name=None):
        if not self.sheetnames:
            raise ValueError("No worksheets in file")

        if name is None:
            name = self.active
        elif isinstance(name, int):
            name = self.sheetnames[name]

        return name

    def _get_cellname(self, row, column):
        row = int(row)
        try:
            column = int(column)
            column = get_column_letter(column)
        except ValueError:
            pass
        return "%s%s" % (column, row)

    def _to_index(self, value):
        value = int(value) if value is not None else 1
        if value < 1:
            raise ValueError("Invalid row index")
        return value

    def create(self):
        self._book = openpyxl.Workbook()
        self._extension = None

    def open(self, path=None, read_only=False, write_only=False, data_only=False):
        self._read_only = read_only
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")

        try:
            extension = pathlib.Path(path).suffix
        except TypeError:
            extension = None

        options = {"filename": path, "data_only": data_only}

        # Only set mode arguments if truthy, otherwise openpyxl complains
        if read_only and write_only:
            raise ValueError("Unable to use both write_only and read_only mode")
        elif read_only:
            options["read_only"] = True
        elif write_only:
            options["write_only"] = True

        if extension in (".xlsm", ".xltm"):
            options["keep_vba"] = True

        self._book = openpyxl.load_workbook(**options)
        self._extension = extension

    def close(self):
        self._book.close()
        self._book = None
        self._extension = None
        self._active = None

    def validate_content(self):
        self._validate_content(self._book.properties)

    def save(self, path=None):
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")

        self._book.save(filename=path)

    def create_worksheet(self, name):
        self._book.create_sheet(title=name)
        self.active = name

    def read_worksheet(self, name=None, header=False, start=None) -> List[dict]:
        name = self._get_sheetname(name)
        sheet = self._book[name]
        start = self._to_index(start)

        if start > sheet.max_row or self.is_sheet_empty(sheet):
            return []

        if header:
            columns = [cell.value for cell in sheet[start]]
            start += 1
        else:
            columns = [get_column_letter(i + 1) for i in range(sheet.max_column)]

        columns = [str(value) if value is not None else value for value in columns]
        columns = ensure_unique(columns)

        data = []
        for cells in sheet.iter_rows(min_row=start):
            row = {}
            for c, acell in enumerate(cells):
                column = columns[c]
                if column is not None:
                    row[column] = acell.value
            data.append(row)

        self.active = name
        return data

    def append_worksheet(
        self,
        name=None,
        content=None,
        header=False,
        start=None,
        formatting_as_empty=False,
    ):
        content = Table(content)
        if not content:
            return

        sheet_name = self._get_sheetname(name)
        sheet = self._book[sheet_name]
        start = self._to_index(start)
        is_empty = self.is_sheet_empty(sheet)

        if header and not is_empty:
            columns = [cell.value for cell in sheet[start]]
        else:
            columns = content.columns

        if header and is_empty:
            sheet.append(columns)

        if formatting_as_empty:
            self._append_on_first_empty_based_on_values(content, columns, sheet)
        else:
            self._default_append_rows(content, columns, sheet)

        self.active = sheet_name

    def _append_on_first_empty_based_on_values(self, content, columns, sheet):
        first_empty_row: Optional[int] = None
        for row_num in range(sheet.max_row, 0, -1):
            if all(cell.value is None for cell in sheet[row_num]):
                first_empty_row = row_num
            else:
                break
        first_empty_row: int = first_empty_row or sheet.max_row + 1
        for row_idx, row in enumerate(content):
            values = self._row_to_values(row, columns)
            for cell_idx, acell in enumerate(sheet[first_empty_row + row_idx]):
                try:
                    acell.value = values[cell_idx]
                except IndexError:
                    pass

    def _default_append_rows(self, content, columns, sheet):
        for row in content:
            values = self._row_to_values(row, columns)
            sheet.append(values)

    def _row_to_values(self, row, columns):
        values = [""] * len(columns)
        for column, value in row.items():
            try:
                index = columns.index(column)
                values[index] = value
            except ValueError:
                pass
        return values

    def remove_worksheet(self, name=None):
        name = self._get_sheetname(name)
        others = [sheet for sheet in self.sheetnames if sheet != name]

        if not others:
            raise ValueError("Workbook must have at least one other worksheet")

        if name == self.active:
            self.active = others[0]

        sheet = self._book[name]
        self._book.remove(sheet)

    def rename_worksheet(self, title, name=None):
        title = str(title)
        name = self._get_sheetname(name)
        sheet = self._book[name]

        sheet.title = title
        self.active = title

    def find_empty_row(self, name=None):
        name = self._get_sheetname(name)
        sheet = self._book[name]

        for idx in reversed(range(sheet.max_row)):
            idx += 1  # Convert to 1-based indexing
            if any(value for value in sheet[idx]):
                return idx + 1  # Return first empty row

        return 1

    def get_cell_value(self, row, column, name=None):
        name = self._get_sheetname(name)
        sheet = self._book[name]
        cell = self._get_cellname(row, column)

        return sheet[cell].value

    def set_cell_value(self, row, column, value, name=None):
        name = self._get_sheetname(name)
        sheet = self._book[name]
        cell = self._get_cellname(row, column)

        sheet[cell].value = value

    def set_cell_format(self, row, column, fmt, name=None):
        name = self._get_sheetname(name)
        sheet = self._book[name]
        cell = self._get_cellname(row, column)

        sheet[cell].number_format = str(fmt)

    def insert_image(self, row, column, image, name=None):
        name = self._get_sheetname(name)
        sheet = self._book[name]
        cell = self._get_cellname(row, column)

        # For compatibility with openpyxl
        stream = BytesIO()
        image.save(stream, format=image.format)
        image.fp = stream

        img = openpyxl.drawing.image.Image(image)
        img.anchor = cell
        sheet.add_image(img)


class XlsWorkbook(BaseWorkbook):
    """Container for manipulating legacy Excel files (.xls)"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._images = []

    @staticmethod
    def is_sheet_empty(sheet):
        return not any([sheet.ncols, sheet.nrows])

    @property
    def sheetnames(self):
        return [sheet.name for sheet in self._book.sheets()]

    @property
    def active(self):
        if not self._active:
            for sheet in self._book.sheets():
                if sheet.sheet_visible:
                    self._active = sheet.name
                    break
            else:
                self._active = self.sheetnames[0]

        return self._active

    @active.setter
    def active(self, value):
        if isinstance(value, int):
            value = self.sheetnames[value]
        elif value not in self.sheetnames:
            raise ValueError(f"Unknown worksheet: {value}")

        for sheet in self._book.sheets():
            match = int(sheet.name == value)
            sheet.sheet_selected = match
            sheet.sheet_visible = match

        self._active = value

    @property
    def extension(self):
        return self._extension

    def _get_sheetname(self, name):
        if self._book.nsheets == 0:
            raise ValueError("No worksheets in file")

        if name is None:
            name = self.active
        elif isinstance(name, int):
            name = self.sheetnames[name]
        elif name not in self.sheetnames:
            raise ValueError(f"Unknown worksheet: {name}")

        return name

    def _get_cell(self, row, column):
        row = int(row)
        try:
            column = int(column)
        except ValueError:
            column = get_column_index(column)
        return row - 1, column - 1

    def _to_index(self, value):
        value = (int(value) - 1) if value is not None else 0
        if value < 0:
            raise ValueError("Invalid row index")
        return value

    def create(self, sheet="Sheet"):
        fd = BytesIO()
        try:
            book = xlwt.Workbook()
            book.add_sheet(sheet)
            book.save(fd)
            fd.seek(0)
            self.open(fd)
        finally:
            fd.close()

        self._extension = None

    def open(self, path=None, read_only=False, write_only=False, data_only=False):
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")

        try:
            extension = pathlib.Path(path).suffix
        except TypeError:
            extension = None

        options = {"on_demand": True, "formatting_info": True}

        if read_only or write_only or data_only:
            self.logger.info(
                "Modes read_only/write_only/data_only not supported with .xls"
            )

        if hasattr(path, "read"):
            options["file_contents"] = path.read()
        else:
            options["filename"] = path

        self._book = xlrd.open_workbook(**options)
        self._extension = extension
        self._images = []

    def close(self):
        self._book.release_resources()
        self._book = None
        self._extension = None
        self._active = None
        self._images = []

    @contextmanager
    def _book_write(self):
        book = xlutils_copy(self._book)
        yield book

        fd = BytesIO()
        try:
            book.save(fd)
            fd.seek(0)
            self.close()
            self.open(fd)
        finally:
            fd.close()

    def validate_content(self):
        self._validate_content(self._book)

    def save(self, path=None):
        path = path or self.path
        if not path:
            raise ValueError("No path defined for workbook")

        book = xlutils_copy(self._book)
        self._insert_images(book)
        book.save(path)

    def create_worksheet(self, name):
        with self._book_write() as book:
            book.add_sheet(name)

        self.active = name

    def read_worksheet(self, name=None, header=False, start=None) -> List[dict]:
        name = self._get_sheetname(name)
        sheet = self._book.sheet_by_name(name)
        start = self._to_index(start)

        if start >= sheet.nrows:
            return []

        if header:
            columns = [self._parse_type(cell) for cell in sheet.row(start)]
            start += 1
        else:
            columns = [get_column_letter(i + 1) for i in range(sheet.ncols)]

        columns = [value if value != "" else None for value in columns]
        columns = [str(value) if value is not None else value for value in columns]
        columns = ensure_unique(columns)

        data = []
        for r in range(start, sheet.nrows):
            row = {}
            for c in range(sheet.ncols):
                column = columns[c]
                if column is not None:
                    cell = sheet.cell(r, c)
                    row[column] = self._parse_type(cell)
            data.append(row)

        self.active = name
        return data

    def _parse_type(self, cell):
        value = cell.value

        if cell.ctype == xlrd.XL_CELL_DATE:
            value = xlrd.xldate_as_datetime(value, self._book.datemode)
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            value = bool(value)
        elif cell.ctype == xlrd.XL_CELL_ERROR:
            value = xlrd.biffh.error_text_from_code.get(value, "#ERROR")
        elif cell.ctype == xlrd.XL_CELL_NUMBER and value.is_integer():
            value = int(value)

        return value

    def append_worksheet(
        self,
        name=None,
        content=None,
        header=False,
        start=None,
        formatting_as_empty=False,
    ):
        content = Table(content)
        if not content:
            return

        name = self._get_sheetname(name)
        sheet_read = self._book.sheet_by_name(name)
        start = self._to_index(start)
        is_empty = self.is_sheet_empty(sheet_read)

        if header and not is_empty:
            columns = [cell.value for cell in sheet_read.row(start)]
        else:
            columns = content.columns

        with self._book_write() as book:
            sheet_write = book.get_sheet(name)
            # TODO. target worksheet cell formatting is overwritten.
            # It would be preferable to preserve formatting in the
            # target worksheet.
            if formatting_as_empty:
                start_row = self._return_first_empty_row(sheet_read)
            else:
                start_row = sheet_read.nrows
            if header and is_empty:
                for column, value in enumerate(columns):
                    sheet_write.write(0, column, value)
                start_row += 1

            for r, row in enumerate(content, start_row):
                for column, value in row.items():
                    sheet_write.write(r, columns.index(column), value)

        self.active = name

    def _return_first_empty_row(self, sheet):
        first_empty_row: Optional[int] = None
        for row_num in range(sheet.nrows - 1, 0, -1):
            if all(cell.value == "" for cell in sheet[row_num]):
                first_empty_row = row_num
            else:
                break
        first_empty_row: int = first_empty_row or sheet.nrows
        return first_empty_row

    def remove_worksheet(self, name=None):
        name = self._get_sheetname(name)
        others = [sheet for sheet in self.sheetnames if sheet != name]

        if not others:
            raise ValueError("Workbook must have at least one other worksheet")

        if name == self.active:
            self.active = others[0]

        with self._book_write() as book:
            # This is pretty ugly, but there seems to be no other way to
            # remove sheets from the xlwt.Workbook instance
            # pylint: disable=protected-access
            book._Workbook__worksheets = [
                sheet for sheet in book._Workbook__worksheets if sheet.name != name
            ]
            book._Workbook__active_sheet = next(
                idx
                for idx, sheet in enumerate(book._Workbook__worksheets)
                if sheet.name == self.active
            )

    def rename_worksheet(self, title, name=None):
        title = str(title)
        name = self._get_sheetname(name)

        with self._book_write() as book:
            sheet = book.get_sheet(name)
            sheet.name = title

        self.active = title

    def find_empty_row(self, name=None):
        name = self._get_sheetname(name)
        sheet = self._book.sheet_by_name(name)

        for row in reversed(range(sheet.nrows)):
            if any(cell.value for cell in sheet.row(row)):
                # Convert to 1-based indexing and
                # return first empty row
                return row + 2

        return 1

    def get_cell_value(self, row, column, name=None):
        name = self._get_sheetname(name)
        sheet = self._book.sheet_by_name(name)
        row, column = self._get_cell(row, column)

        return sheet.cell_value(row, column)

    def set_cell_value(self, row, column, value, name=None):
        name = self._get_sheetname(name)
        row, column = self._get_cell(row, column)

        with self._book_write() as book:
            sheet = book.get_sheet(name)
            sheet.write(row, column, value)

    def set_cell_format(self, row, column, fmt, name=None):
        name = self._get_sheetname(name)
        sheet = self._book.sheet_by_name(name)
        row, column = self._get_cell(row, column)

        value = sheet.cell_value(row, column)
        style = xlwt.XFStyle()
        style.num_format_str = str(fmt)

        with self._book_write() as book:
            sheet = book.get_sheet(name)
            sheet.write(row, column, value, style)

    def insert_image(self, row, column, image, name=None):
        name = self._get_sheetname(name)
        row, column = self._get_cell(row, column)
        self._images.append((name, row, column, image))

    def _insert_images(self, book):
        for name, row, column, image in self._images:
            stream = BytesIO()
            image.save(stream, format="BMP")
            bitmap = stream.getvalue()

            sheet = book.get_sheet(name)
            sheet.insert_bitmap_data(bitmap, row, column)
